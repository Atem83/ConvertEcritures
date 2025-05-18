import re

import polars as pl
from openpyxl import load_workbook

from py_convert.error import run_error
from py_convert.format_import import ImportBase

class ImportLabo(ImportBase):
    """Gestion d'import d'un fichier Excel au format LABO."""
    
    def name(self):
        return "LABO"
    
    def validate_format(self):
        if self.path.suffix.lower() == ".xls":
            return True
        elif self.path.suffix.lower() == ".xlsx":
            return True
        else:
            run_error(f"Le format {self.name()} nécessite un fichier .xls")
            return False
    
    def process_file(self):
        df = pl.read_excel(
            source=self.path, 
            engine="calamine", 
            read_options={
                "column_names": self.get_titles(), 
                "skip_rows": 3
                }
            )
        date = self.get_date(df, "Date de vente")

        # Conserve uniquement les colonnes pertinentes
        df = df.select(
            pl.col("Type de vente").cast(pl.String), 
            pl.col(
                *[col for col in df.columns if col.startswith("H.T.")]
                ).cast(pl.Float64), 
            pl.col(
                *[col for col in df.columns if col.startswith("TVA")]
                ).cast(pl.Float64)
            )

        # Conserve uniquement les données pertinentes
        df = df.filter(pl.col("Type de vente").str.starts_with("Total :"))

        # Reformate le nom des pays
        df = df.with_columns(pl.col("Type de vente").str.replace("Total : ", ""))

        # Création de l'écriture comptable
        entries = []

        # Ajout du compte client
        entries.append({
            "JournalCode": "VE",
            "EcritureDate": date,
            "CompteNum": "41100000",
            "CompAuxNum": "CCLIENT",
            "PieceDate": date,
            "EcritureLib": "CLIENTS B TO C", 
            "Debit": 0.0, 
            "Credit": 0.0,
            })

        for country in df.to_dicts():
            name = country["Type de vente"].upper()

            # Ajout des comptes de produits et de TVA
            for key, value in country.items():
                percent = self.get_vat(key)
                if percent is None or value == 0:
                    continue
                
                label = name + " " + percent + " " + date.strftime("%m/%Y")

                if key.startswith("H.T."):
                    account = "70600000"
                elif key.startswith("TVA"):
                    account = "44571000"
                else:
                    account = None

                if value > 0:
                    debit = 0.0
                    credit = value
                else:
                    debit = -value
                    credit = 0.0

                entries.append({
                    "JournalCode": "VE",
                    "EcritureDate": date,
                    "CompteNum": account,
                    "CompAuxNum": None,
                    "PieceDate": date,
                    "EcritureLib": label, 
                    "Debit": debit, 
                    "Credit": credit,
                    })

        df = pl.DataFrame(entries)
        df = df.with_columns(
            pl.when(pl.col("CompteNum").str.starts_with("411"))
            .then(df["Credit"].sum() - df["Debit"].sum())
            .otherwise(pl.col("Debit"))
            .alias("Debit")
            )
        df = df.sort(
            ["CompAuxNum", "EcritureLib", "CompteNum"],
            descending=[False, False, True],
            nulls_last=[True, False, False]
            )
        
        return df
    
    def get_titles(self):
        """Récupère les en-têtes des colonnes"""
        wb = load_workbook(filename=self.path)
        ws = wb.worksheets[0]
        end_col = ws.max_column
        titles = []
        
        for col in range(1, end_col):
            cell1 = str(ws.cell(1, col).value or '')
            cell3 = str(ws.cell(3, col).value or '')
            if cell3.strip() != "TVA":
                cell2 = self.get_vat(str(ws.cell(2, col).value or ''))

            if cell1 is None :
                cell1 = ""
            else:
                cell1 = cell1.strip()

            if cell2 is None :
                cell2 = ""
            else:
                cell2 = cell2.strip()

            if cell3 is None :
                cell3 = ""
            else:
                cell3 = cell3.strip()

            name = f"{cell3} {cell2} {cell1}".strip()
            titles.append(name)
        
        return titles
    
    def get_vat(self,text: str):
        """Récupère le % de TVA d'un champ texte"""
        text = text.replace(',', '.')
        match = re.search(r'(\d+\.?\d*)\s*%', str(text))
        if match:
            return match.group(1) + "%"
        return None

    def get_date(self, df: pl.DataFrame, col_name: str) -> pl.Date:
        """Récupère le mois et l'année le plus rencontré d'un DataFrame"""
        df = df.select(pl.col(col_name).cast(pl.Date))
        df = df.filter(pl.col(col_name).is_not_null())

        # Compte le nombre d'occurence de chaque groupe année/mois
        df = (df.with_columns([
                pl.col(col_name).dt.year().alias("Année"),
                pl.col(col_name).dt.month().alias("Mois")
            ]).group_by(["Année", "Mois"])
              .agg(pl.len().alias("Count"))
              .sort(["Année", "Mois"])
              )

        # Conserve uniquement le groupe année/mois le plus rencontré
        df = df.filter(pl.col("Count").max() == pl.col("Count"))

        # Récupère le dernier jour du mois et transforme la valeur en date
        df = df.with_columns(
            pl.date(
                pl.col("Année"),
                pl.col("Mois"),
                1
                ).dt.month_end().alias("Date")
            )

        return df["Date"][0]